#!/usr/bin/env python3
"""
build_payroll_with_scheduled.py

Join a Toast payroll export with a Sling shifts export to produce a per-(employee,
job-title, station) CSV that has scheduled hours, expected pay, actual hours, and
actual pay on each row -- plus bucket subtotals and a grand total at the bottom.

As of the tag-driven version, each BOH shift in the Sling export carries a station
tag after a " . " separator (e.g. "BOH . Mainline", "BOH . An SE"). These tags
drive the BOH buckets directly, so AN BOH.txt is no longer needed.

Inputs
------
By default, the script looks for these files in the script's own folder:
  payroll_export.csv          Toast payroll export
  shifts-export.xls           Sling shifts export (.xls extension but actually XLSX)
  SALARIED EMPLOYEES.txt      One name per line; "Name - Title" lines allowed
  Overnight_Employees.txt     Sections: FOH / SL / BOH, names underneath

Each input path can be overridden with the corresponding CLI flag.

  --payroll        Toast payroll export CSV
  --shifts         Sling shifts export
  --salaried       SALARIED EMPLOYEES.txt
  --overnight      Overnight_Employees.txt
  --out            Detail CSV path. Default: <script_folder>/payroll_with_scheduled_employees.csv
  --out-buckets    Buckets CSV path. Default: <script_folder>/payroll_with_scheduled_buckets.csv
  --rp-sales       RP Net Adj Sales (optional; skips the interactive prompt)
  --an-sales       AN Net Adj Sales (optional; skips the interactive prompt)

Outputs
-------
Two CSVs:
  1. Detail CSV  -- one row per (employee, job-title, station), no subtotals.
  2. Buckets CSV -- the bucket subtotals, plus two columns expressing each
     bucket's Total Pay as a labor % of sales. At the end of the run the script
     prompts for "RP Net Adj Sales = " and "AN Net Adj Sales = " (each parsed
     leniently from $ / comma formatting); AN buckets are measured against AN
     sales, RP buckets against RP sales, the TOTAL against combined, and every
     bucket also against combined AN+RP sales.

Buckets (see buckets.txt)
-------------------------
  AN FOH
  AN BOH                  <- Sling tag "An SE" on a BOH-title shift
  RP FOH
  RP BOH - Kitchen Prep   <- Sling tag "Kitchen Prep" (also "Prep")
  RP BOH - Grill          <- Sling tag "Grill"
  RP BOH - Mainline       <- Sling tag "Mainline"
  RP BOH - Dishwasher     <- Sling tag "Diswasher"/"Dishwasher"
  RP BOH - SL + Other     <- BOH SL and any untagged BOH-title shift

What it does
------------
1. Parses the shifts XLSX and aggregates scheduled hours per
   (employee, payroll job title, station tag). The Sling role string is split on
   the first " . " separator into a base position (matches the payroll Job Title)
   and a station tag.
2. Loads payroll, drops columns Sebastian doesn't want (tips/IDs/location/etc.),
   excludes salaried rows (Job Title starts with "Salaried" / equals
   "Kitchen Manager", or Hourly Rate <= 0 / NaN) and Training rows.
3. Joins scheduled hours onto each payroll row, using the exact (employee, job
   title) key with a last-name fallback for Sling/Toast name-order mismatches.
4. Computes Expected Pay = Hourly Rate * Scheduled Hours, and Pay Variance =
   Total Pay - Expected Pay (positive = over scheduled, negative = under).
5. EXPLODES each BOH-title payroll row into one row per scheduled station tag,
   allocating that row's actual Regular/Overtime hours and pay across the
   stations IN PROPORTION to the station's share of scheduled hours. Toast only
   reports actuals at the BOH-title level (no station tag), so this proportional
   split is how per-station actual cost is derived. Rounding residuals land on
   the largest-share station so every column reconciles exactly to the source
   row -- grand totals are unchanged. Untagged BOH-title rows (incl. BOH SL) and
   rows with no Sling match collapse into a single "SL + Other" station.
6. Buckets each row using the station tag, with Overnight_Employees.txt routing
   Overnight / Overnight SL rows to the FOH or BOH side.
7. Appends subtotal rows in this order:
       AN LABOR (AN FOH + AN BOH)
       AN FOH
       AN BOH
       RP LABOR (RP BOH + RP FOH)
       RP BOH
         RP BOH - Kitchen Prep
         RP BOH - Grill
         RP BOH - Mainline
         RP BOH - Dishwasher
         RP BOH - SL + Other
       RP FOH
       TOTAL

Usage
-----
    python3 build_payroll_with_scheduled.py
    python3 build_payroll_with_scheduled.py --payroll path/to/payroll.csv --out /tmp/out.csv
"""

import argparse
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: pip install openpyxl\n")
    sys.exit(2)


BULLET = "\u2022"            # the " . " separator Sling uses is actually " <bullet> "
SEP = f" {BULLET} "


# ---------------------------------------------------------------------------
# Money input parsing
# ---------------------------------------------------------------------------

def parse_money(text: str) -> float:
    """Parse a loosely formatted dollar value into a float.

    Handles things like '$48,231.50', '48231.5', '  $1,200 ', '(500)' (negative),
    and a trailing 'k'/'m' multiplier. Raises ValueError if nothing numeric is left.
    """
    s = str(text).strip()
    if not s:
        raise ValueError("empty value")
    neg = False
    if s.startswith("(") and s.endswith(")"):   # accounting-style negative
        neg, s = True, s[1:-1]
    mult = 1.0
    low = s.lower()
    if low.endswith("k"):
        mult, s = 1_000.0, s[:-1]
    elif low.endswith("m"):
        mult, s = 1_000_000.0, s[:-1]
    # Strip currency symbols, commas, spaces -- keep digits, dot, sign.
    cleaned = re.sub(r"[^0-9.\-]", "", s)
    if cleaned in ("", "-", ".", "-.", "+"):
        raise ValueError(f"could not parse a number from {text!r}")
    val = float(cleaned) * mult
    return -val if neg else val


def prompt_sales(label: str, provided=None) -> float:
    """Return a sales figure: use `provided` if given, else prompt until valid."""
    if provided is not None:
        return parse_money(provided)
    while True:
        try:
            return parse_money(input(label))
        except (ValueError, EOFError) as e:
            sys.stderr.write(f"  '{e}' -- please enter a dollar amount (e.g. 48,231.50)\n")


# ---------------------------------------------------------------------------
# Shifts XLSX parser (self-contained; mirrors the project's shifts-xlsx-parser)
# ---------------------------------------------------------------------------

# Matches "9h 30 min", "9h", "45 min", etc.
DURATION_RE = re.compile(r"(\d+)h(?:\s*(\d+)\s*min)?|(\d+)\s*min")


def parse_duration_to_hours(s: str) -> float:
    """Extract hours as a float from a duration string."""
    m = DURATION_RE.search(s or "")
    if not m:
        return 0.0
    if m.group(1):
        return int(m.group(1)) + (int(m.group(2)) / 60.0 if m.group(2) else 0.0)
    return int(m.group(3)) / 60.0


def parse_cell(text):
    """Split one day's shifts cell into [(name, hours, role), ...].

    Sling cells contain multiple employee blocks separated by blank lines.
    Each block is typically 3 lines (name / time-range-and-duration / role),
    though sometimes 2 (name / time-range-and-duration) when role is missing.
    """
    if not text:
        return []
    blocks = re.split(r"\n\s*\n", str(text).strip())
    out = []
    for block in blocks:
        lines = [ln.strip() for ln in block.split("\n") if ln.strip()]
        if len(lines) < 2:
            continue
        name = lines[0]
        # Skip Sling's "Unassigned" / "Available" pseudo-employees.
        if name.lower().startswith(("unassigned", "available")):
            continue
        hours = parse_duration_to_hours(lines[1])
        role = lines[2] if len(lines) > 2 else ""
        out.append((name, hours, role))
    return out


def open_shifts_workbook(path: Path):
    """The shifts export has a .xls extension but is actually XLSX inside.

    openpyxl rejects files based on extension, so copy to a temp .xlsx if needed.
    Caller is responsible for not deleting the temp until it's done reading.
    """
    head = path.read_bytes()[:4]
    if head != b"PK\x03\x04":
        raise ValueError(
            f"{path} doesn't look like XLSX (header={head!r}). "
            f"The shifts export should be a ZIP-based XLSX file."
        )
    if path.suffix.lower() == ".xlsx":
        return load_workbook(str(path), data_only=True), None
    tmpdir = tempfile.mkdtemp(prefix="shifts_")
    tmp_path = Path(tmpdir) / "shifts.xlsx"
    shutil.copy(path, tmp_path)
    return load_workbook(str(tmp_path), data_only=True), tmpdir


def split_sling_role(role: str):
    """'BOH . Mainline' -> ('BOH', 'Mainline'); 'BOH' -> ('BOH', '').

    Sling roles concatenate a base position with an optional station/location tag
    using ' <bullet> ' as the separator. The base position before the first
    separator matches the payroll Job Title verbatim; the remainder is the tag.
    """
    parts = (role or "").split(SEP, 1)
    base = parts[0].strip()
    tag = parts[1].strip() if len(parts) > 1 else ""
    return base, tag


def aggregate_scheduled_hours(shifts_path: Path) -> dict:
    """Return {(employee_first_last, payroll_job_title): {station_tag: hours}}.

    Reads row 7 of the shifts XLSX (the data row containing per-day cells)
    and aggregates across the day columns.
    """
    wb, tmpdir = open_shifts_workbook(shifts_path)
    try:
        ws = wb.active
        data_row = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
        sched = defaultdict(lambda: defaultdict(float))
        for cell in data_row[:8]:
            if not cell:
                continue
            for name, hours, role in parse_cell(cell):
                base, tag = split_sling_role(role)
                sched[(name, base)][tag] += hours
        # convert to plain dicts
        return {k: dict(v) for k, v in sched.items()}
    finally:
        if tmpdir:
            shutil.rmtree(tmpdir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Name normalization
# ---------------------------------------------------------------------------

# Project memory: some employees appear under slightly different names across
# Sling and Toast. Map the Sling spelling to the payroll spelling here.
SLING_TO_PAYROLL_NAME_ALIASES = {
    "Veronica Cruz Cruz De Vasquez": "Veronica Cruz De Vasquez",
}


def payroll_to_first_last(name: str) -> str:
    """'Last, First [Middle]' -> 'First Last' (preserving middles)."""
    name = str(name).strip()
    if "," in name:
        last, rest = name.split(",", 1)
        return f"{rest.strip()} {last.strip()}"
    return name


def last_name_token(name: str) -> str:
    """Lowercased last-name token for fuzzy fallback matching."""
    s = str(name).strip()
    if not s:
        return ""
    if "," in s:
        return s.split(",")[0].strip().split()[-1].lower()
    return s.split()[-1].lower()


# ---------------------------------------------------------------------------
# Roster files
# ---------------------------------------------------------------------------

def load_overnight_roster(path: Path) -> dict:
    """Return {last_name_key: 'FOH' | 'BOH' | 'SL'}.

    File format: section headers 'FOH', 'SL', 'BOH' followed by names underneath.
    SL is treated as BOH-side for bucketing per project convention.
    """
    if not path or not path.exists():
        return {}
    out = {}
    section = None
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or line.lower().startswith("overnight_"):
            continue
        if line.upper() in ("FOH", "BOH", "SL"):
            section = line.upper()
            continue
        if section:
            out[last_name_token(line)] = section
    return out


# ---------------------------------------------------------------------------
# Bucketing
# ---------------------------------------------------------------------------

RP_FOH_TITLES = {"FOH", "FOH (Late night)", "Shift Lead", "Shift Lead (Late night)"}
RP_BOH_TITLES = {"BOH", "BOH (Late night)", "BOH SL", "BOH SL (Late Night)"}
AN_FOH_TITLES = {"An FOH", "An Line Lead"}
OVERNIGHT_TITLES = {"Overnight", "Overnight SL"}

# Granular RP BOH station buckets and the order they print in.
RP_BOH_STATIONS = [
    "RP BOH - Kitchen Prep",
    "RP BOH - Grill",
    "RP BOH - Mainline",
    "RP BOH - Dishwasher",
    "RP BOH - SL + Other",
]


def station_bucket_for_tag(tag: str) -> str:
    """Map a BOH-side station tag to its granular bucket label."""
    t = (tag or "").strip().lower()
    if t in ("an se",):
        return "AN BOH"
    if t in ("kitchen prep", "prep"):
        return "RP BOH - Kitchen Prep"
    if t == "grill":
        return "RP BOH - Grill"
    if t == "mainline":
        return "RP BOH - Mainline"
    if t in ("diswasher", "dishwasher"):
        return "RP BOH - Dishwasher"
    # Untagged BOH, BOH SL, or anything unrecognized -> SL + Other
    return "RP BOH - SL + Other"


def row_side(job_title, employee, overnight_roster):
    """Return one of 'AN FOH', 'RP FOH', 'BOH' (explode), or 'OTHER'."""
    if job_title in AN_FOH_TITLES:
        return "AN FOH"
    if job_title in RP_FOH_TITLES:
        return "RP FOH"
    if job_title in RP_BOH_TITLES:
        return "BOH"
    if job_title in OVERNIGHT_TITLES:
        side = overnight_roster.get(last_name_token(employee))
        return "RP FOH" if side == "FOH" else "BOH"
    return "OTHER"


def parent_bucket(bucket: str) -> str:
    """Roll a granular bucket up to its parent (AN FOH / AN BOH / RP FOH / RP BOH)."""
    if bucket.startswith("RP BOH"):
        return "RP BOH"
    return bucket


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

NUMERIC_OUTPUT_COLS = [
    "Hourly Rate", "Scheduled Hours", "Expected Pay",
    "Regular Hours", "Overtime Hours",
    "Regular Pay", "Overtime Pay", "Total Pay",
    "Pay Variance",
]

FINAL_COLUMN_ORDER = [
    "Employee", "Job Title", "Hourly Rate",
    "Scheduled Hours", "Expected Pay",
    "Regular Hours", "Overtime Hours",
    "Regular Pay", "Overtime Pay", "Total Pay",
    "Pay Variance",
]

# Actual columns split proportionally when a BOH row is exploded by station.
ALLOC_COLS = ["Regular Hours", "Overtime Hours", "Regular Pay", "Overtime Pay", "Total Pay"]


def allocate_proportional(totals: dict, weights: list):
    """Split each total across weighted buckets, residual to the largest weight.

    totals:  {col_name: total_value}
    weights: list of (key, weight) with weights summing to ~1.0
    Returns: {key: {col_name: value}} reconciling exactly to totals.
    """
    keys = [k for k, _ in weights]
    # Largest-weight key absorbs rounding residual.
    residual_key = max(weights, key=lambda kw: kw[1])[0] if weights else None
    alloc = {k: {} for k in keys}
    for col, total in totals.items():
        running = 0.0
        for k, w in weights:
            if k == residual_key:
                continue
            v = round(total * w, 2)
            alloc[k][col] = v
            running += v
        if residual_key is not None:
            alloc[residual_key][col] = round(total - running, 2)
    return alloc


def build(payroll_path, shifts_path, salaried_path, overnight_path):
    # 1) Parse shifts -> {(sling_name, base_title): {tag: hours}}
    sched_raw = aggregate_scheduled_hours(Path(shifts_path))

    # Apply name aliases on the Sling side so the join key matches payroll.
    by_name = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for (name, base), tagmap in sched_raw.items():
        name = SLING_TO_PAYROLL_NAME_ALIASES.get(name, name)
        for tag, hrs in tagmap.items():
            by_name[name][base][tag] += hrs
    # collapse to plain dicts: name -> base -> {tag: hrs}
    by_name = {n: {b: dict(tm) for b, tm in bm.items()} for n, bm in by_name.items()}

    # Last-name fallback index.
    last_to_names = defaultdict(set)
    for sling_name in by_name:
        last_to_names[last_name_token(sling_name)].add(sling_name)

    def lookup_sched_tags(sling_name, job_title) -> dict:
        """Return {tag: hours} scheduled for (name, job_title), or {}."""
        if job_title in by_name.get(sling_name, {}):
            return by_name[sling_name][job_title]
        # Fallback: last-name match, only when the payroll name isn't otherwise
        # in Sling and exactly one Sling person shares the last-name token.
        if sling_name in by_name:
            return {}
        candidates = last_to_names.get(last_name_token(sling_name), set())
        if len(candidates) == 1:
            (cand,) = candidates
            if job_title in by_name[cand]:
                return by_name[cand][job_title]
        elif len(candidates) > 1:
            sys.stderr.write(
                f"WARN: ambiguous last-name match for payroll {sling_name!r} "
                f"({job_title!r}); Sling has {sorted(candidates)}; using 0 scheduled hours\n"
            )
        return {}

    # 2) Load payroll, drop unused columns, exclude salaried + training rows.
    df = pd.read_csv(payroll_path)
    drop_cols = [
        "Net Sales", "Declared Tips", "Non-Cash Tips", "Total Tips",
        "Tips Withheld", "Total Gratuity", "Employee ID", "Job Code",
        "Location", "Location Code",
    ]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns])

    salaried_title_mask = df["Job Title"].str.startswith("Salaried") | (df["Job Title"] == "Kitchen Manager")
    training_mask = df["Job Title"] == "Training"
    zero_rate_mask = df["Hourly Rate"].fillna(0) <= 0
    df = df[~(salaried_title_mask | training_mask | zero_rate_mask)].copy()

    for c in ["Hourly Rate", "Regular Hours", "Overtime Hours", "Regular Pay", "Overtime Pay", "Total Pay"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    overnight = load_overnight_roster(Path(overnight_path)) if overnight_path else {}

    # 3) Build output rows: single row for FOH-side, exploded rows for BOH-side.
    out_rows = []
    for _, row in df.iterrows():
        emp = row["Employee"]
        jt = row["Job Title"]
        rate = row["Hourly Rate"]
        sling_name = payroll_to_first_last(emp)
        tagmap = lookup_sched_tags(sling_name, jt)
        side = row_side(jt, emp, overnight)

        if side in ("AN FOH", "RP FOH", "OTHER"):
            sched_total = round(sum(tagmap.values()), 2)
            out_rows.append({
                "Employee": emp, "Job Title": jt, "Hourly Rate": rate,
                "Scheduled Hours": sched_total,
                "Expected Pay": round(rate * sched_total, 2),
                "Regular Hours": row["Regular Hours"], "Overtime Hours": row["Overtime Hours"],
                "Regular Pay": row["Regular Pay"], "Overtime Pay": row["Overtime Pay"],
                "Total Pay": row["Total Pay"],
                "_bucket": side,
            })
            if side == "OTHER":
                sys.stderr.write(f"WARN: unbucketed row {emp!r} / {jt!r}\n")
            continue

        # BOH-side: explode by station tag, allocating actuals proportionally.
        tags = {t: h for t, h in tagmap.items() if h > 0}
        if not tags:
            # No scheduled station info: collapse to one SL + Other row.
            out_rows.append({
                "Employee": emp, "Job Title": jt, "Hourly Rate": rate,
                "Scheduled Hours": 0.0, "Expected Pay": 0.0,
                "Regular Hours": row["Regular Hours"], "Overtime Hours": row["Overtime Hours"],
                "Regular Pay": row["Regular Pay"], "Overtime Pay": row["Overtime Pay"],
                "Total Pay": row["Total Pay"],
                "_bucket": "RP BOH - SL + Other",
            })
            continue

        total_sched = sum(tags.values())
        # Stable order: largest scheduled share last so it absorbs the residual.
        ordered = sorted(tags.items(), key=lambda kv: (kv[1], kv[0]))
        weights = [(t, h / total_sched) for t, h in ordered]
        totals = {c: float(row[c]) for c in ALLOC_COLS}
        alloc = allocate_proportional(totals, weights)

        for tag, hrs in ordered:
            a = alloc[tag]
            bucket = station_bucket_for_tag(tag)
            jt_label = jt if not tag else f"{jt}{SEP}{tag}"
            sched_h = round(hrs, 2)
            out_rows.append({
                "Employee": emp, "Job Title": jt_label, "Hourly Rate": rate,
                "Scheduled Hours": sched_h,
                "Expected Pay": round(rate * sched_h, 2),
                "Regular Hours": a["Regular Hours"], "Overtime Hours": a["Overtime Hours"],
                "Regular Pay": a["Regular Pay"], "Overtime Pay": a["Overtime Pay"],
                "Total Pay": a["Total Pay"],
                "_bucket": bucket,
            })

    out = pd.DataFrame(out_rows)
    out["Pay Variance"] = (out["Total Pay"] - out["Expected Pay"]).round(2)
    for c in NUMERIC_OUTPUT_COLS:
        out[c] = out[c].round(2)

    out = out.sort_values(by=["Employee", "Job Title"], kind="stable").reset_index(drop=True)

    # Return detail rows carrying the internal "_bucket" label. Subtotals (with
    # labor-% of sales) are built separately by build_buckets() so they can land
    # in their own CSV.
    return out[FINAL_COLUMN_ORDER + ["_bucket"]]


# ---------------------------------------------------------------------------
# Bucket subtotals CSV (with labor % of sales)
# ---------------------------------------------------------------------------

# Per-bucket concept -> which sales figure is the "own concept" denominator.
#   "AN"       -> AN Net Adj Sales
#   "RP"       -> RP Net Adj Sales
#   "COMBINED" -> AN + RP (used by the grand TOTAL row)
PCT_AN_RP_COL = "% of AN/RP sales"
PCT_COMBINED_COL = "% of Combined AN+RP sales"
BUCKET_COLUMN_ORDER = FINAL_COLUMN_ORDER + [PCT_AN_RP_COL, PCT_COMBINED_COL]


def build_buckets(detail, rp_sales, an_sales):
    """Build the bucket-subtotal DataFrame from detail rows + sales inputs.

    Each bucket's Total Pay is expressed as a % of its own concept's sales
    (AN buckets vs AN sales, RP buckets vs RP sales, TOTAL vs combined) and as a
    % of combined AN+RP sales.
    """
    combined = (rp_sales or 0) + (an_sales or 0)
    bucket_col = detail["_bucket"]
    parent = bucket_col.apply(parent_bucket)

    def subtotal(label, mask, concept):
        s = detail[mask]
        return {
            "Employee": label, "Job Title": "", "Hourly Rate": "",
            "Scheduled Hours": round(s["Scheduled Hours"].sum(), 2),
            "Expected Pay":    round(s["Expected Pay"].sum(), 2),
            "Regular Hours":   round(s["Regular Hours"].sum(), 2),
            "Overtime Hours":  round(s["Overtime Hours"].sum(), 2),
            "Regular Pay":     round(s["Regular Pay"].sum(), 2),
            "Overtime Pay":    round(s["Overtime Pay"].sum(), 2),
            "Total Pay":       round(s["Total Pay"].sum(), 2),
            "Pay Variance":    round(s["Pay Variance"].sum(), 2),
            "_concept":        concept,
        }

    an_foh = parent == "AN FOH"
    an_boh = parent == "AN BOH"
    rp_foh = parent == "RP FOH"
    rp_boh = parent == "RP BOH"

    rows = [
        subtotal("AN LABOR (AN FOH + AN BOH)", an_foh | an_boh, "AN"),
        subtotal("AN FOH",                     an_foh,          "AN"),
        subtotal("AN BOH",                     an_boh,          "AN"),
        subtotal("RP LABOR (RP FOH + RP BOH)", rp_boh | rp_foh, "RP"),
        subtotal("RP FOH",                     rp_foh,          "RP"),
        subtotal("RP BOH",                     rp_boh,          "RP"),
    ]
    for station in RP_BOH_STATIONS:
        rows.append(subtotal(f"  {station}", bucket_col == station, "RP"))
    rows += [
        subtotal("TOTAL",  pd.Series([True] * len(detail), index=detail.index), "COMBINED"),
    ]

    own_denom = {"AN": an_sales, "RP": rp_sales, "COMBINED": combined}

    def pct(amount, denom):
        if not denom:
            return ""
        return round(amount / denom * 100, 2)

    for r in rows:
        concept = r.pop("_concept")
        r[PCT_AN_RP_COL] = pct(r["Total Pay"], own_denom[concept])
        r[PCT_COMBINED_COL] = pct(r["Total Pay"], combined)

    return pd.DataFrame(rows)[BUCKET_COLUMN_ORDER]


def main():
    script_dir = Path(__file__).resolve().parent

    defaults = {
        "payroll":   script_dir / "payroll_export.csv",
        "shifts":    script_dir / "shifts-export.xls",
        "salaried":  script_dir / "SALARIED EMPLOYEES.txt",
        "overnight": script_dir / "Overnight_Employees.txt",
        "out":       script_dir / "payroll_with_scheduled_employees.csv",
        "out_buckets": script_dir / "payroll_with_scheduled_buckets.csv",
    }

    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--payroll",   default=str(defaults["payroll"]),   help=f"Toast payroll export CSV (default: {defaults['payroll'].name})")
    ap.add_argument("--shifts",    default=str(defaults["shifts"]),    help=f"Sling shifts export (default: {defaults['shifts'].name})")
    ap.add_argument("--salaried",  default=str(defaults["salaried"]),  help=f"Salaried employees list (default: {defaults['salaried'].name})")
    ap.add_argument("--overnight", default=str(defaults["overnight"]), help=f"Overnight employees list (default: {defaults['overnight'].name})")
    ap.add_argument("--out",       default=str(defaults["out"]),       help=f"Detail CSV path (default: {defaults['out'].name})")
    ap.add_argument("--out-buckets", default=str(defaults["out_buckets"]), help=f"Buckets CSV path (default: {defaults['out_buckets'].name})")
    ap.add_argument("--rp-sales",  default=None, help="RP Net Adj Sales (skips the interactive prompt if given)")
    ap.add_argument("--an-sales",  default=None, help="AN Net Adj Sales (skips the interactive prompt if given)")
    args = ap.parse_args()

    missing = []
    for label, path in [("payroll", args.payroll), ("shifts", args.shifts),
                        ("salaried", args.salaried), ("overnight", args.overnight)]:
        if not Path(path).exists():
            missing.append(f"  --{label}: {path}")
    if missing:
        sys.stderr.write("ERROR: required input file(s) not found:\n" + "\n".join(missing) + "\n")
        sys.stderr.write(f"\nLooked relative to script folder: {script_dir}\n")
        sys.exit(1)

    # Note: salaried_path is required for argument symmetry, but the salaried
    # exclusion relies on payroll fields (Job Title and Hourly Rate) which catch
    # every salaried row in practice.

    detail = build(
        payroll_path=args.payroll,
        shifts_path=args.shifts,
        salaried_path=args.salaried,
        overnight_path=args.overnight,
    )

    # 1) Detail CSV (per-employee rows only; no subtotals, no internal bucket col)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    detail[FINAL_COLUMN_ORDER].to_csv(out_path, index=False)
    sys.stderr.write(f"wrote {out_path} ({len(detail)} detail rows)\n")

    # 2) Sales inputs (prompted at the end unless supplied via flags)
    rp_sales = prompt_sales("RP Net Adj Sales = ", args.rp_sales)
    an_sales = prompt_sales("AN Net Adj Sales = ", args.an_sales)

    # 3) Buckets CSV (subtotals + labor % of sales)
    buckets = build_buckets(detail, rp_sales=rp_sales, an_sales=an_sales)
    buckets_path = Path(args.out_buckets)
    buckets_path.parent.mkdir(parents=True, exist_ok=True)
    buckets.to_csv(buckets_path, index=False)
    sys.stderr.write(
        f"wrote {buckets_path} ({len(buckets)} bucket rows; "
        f"RP={rp_sales:,.2f} AN={an_sales:,.2f} combined={rp_sales+an_sales:,.2f})\n"
    )


if __name__ == "__main__":
    main()
